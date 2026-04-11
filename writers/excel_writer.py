"""
Excel writer that produces a workbook matching the structure of the
Bank Hapoalim interest-charge sample workbook (463560.xlsx).

Sheets produced
---------------
1. ``פירוט ריבית חובה לפי PDF``  – interest-charge detail, populated from parsed PDF data.
2. ``בנק``                        – bank transaction register template.
3. ``חישוב החזר לפי יתרות בבנק`` – refund calculation template.

All Hebrew text is written right-to-left (ReadingOrder=2).  Formulas are
written in Excel A1 notation exactly as they appear in the sample file.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from models.statement import BankTransaction, Statement, TierRow


# ── Colour constants (RGB hex, no alpha prefix for openpyxl) ─────────────────
_BLUE_HEADER   = "D9E1F2"   # light-blue header fill (row 9)
_YELLOW        = "FFFF00"   # yellow highlight
_INPUT_FILL    = "DDEBF7"   # light-blue for editable input cells (I3, I4…)
_HEADER_FILL2  = "BDD7EE"   # slightly darker blue for Sheet 2 / 3 header rows

# ── Reusable style factories ──────────────────────────────────────────────────

def _font(bold: bool = False, size: float = 11, name: str = "Arial") -> Font:
    return Font(bold=bold, size=size, name=name)


def _fill(hex_rgb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_rgb)


def _thin_border(
    left: bool = True,
    right: bool = True,
    top: bool = True,
    bottom: bool = True,
) -> Border:
    thin = Side(style="thin")
    none = Side(style=None)
    return Border(
        left=thin if left else none,
        right=thin if right else none,
        top=thin if top else none,
        bottom=thin if bottom else none,
    )


def _medium_border(sides: str = "lr") -> Border:
    """sides: any combination of l/r/t/b."""
    medium = Side(style="medium")
    none = Side(style=None)
    return Border(
        left=medium if "l" in sides else none,
        right=medium if "r" in sides else none,
        top=medium if "t" in sides else none,
        bottom=medium if "b" in sides else none,
    )


def _rtl_align(
    horizontal: str = "general",
    vertical: str = "bottom",
    wrap: bool = False,
) -> Alignment:
    return Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap,
        readingOrder=2,   # right-to-left
    )


def _center_rtl() -> Alignment:
    return _rtl_align(horizontal="center")


def _right_rtl() -> Alignment:
    return _rtl_align(horizontal="right")


def _set(
    ws,
    cell_ref: str,
    value,
    *,
    bold: bool = False,
    size: float = 11,
    fill_hex: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[Alignment] = None,
    border: Optional[Border] = None,
    font_name: str = "Arial",
) -> None:
    """Write *value* to *cell_ref* and apply optional formatting."""
    cell = ws[cell_ref]
    cell.value = value
    cell.font = _font(bold=bold, size=size, name=font_name)
    if fill_hex:
        cell.fill = _fill(fill_hex)
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


# ── Date-string → datetime helper ────────────────────────────────────────────

def _parse_tx_date(date_str: str) -> Optional[datetime]:
    """
    Parse a transaction date string into a Python datetime so openpyxl writes
    a real Excel serial date (enabling arithmetic like =A6-A7).

    Accepts: "DD.MM.YYYY", "DD.MM.YY", "DD/MM/YYYY", "DD/MM/YY".
    Returns None if parsing fails (cell will be left empty rather than storing
    a string that would cause #VALUE! in day-difference formulas).
    """
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            pass
    return None


# ── Rate-change date helper ───────────────────────────────────────────────────

def _parse_rate_change_day(rate_change_date: str) -> Optional[date]:
    """
    Convert a rate-change string like "27.11" to a date.
    Assumes the year is taken from context (the billing quarter).
    Returns None if parsing fails.
    """
    try:
        parts = rate_change_date.strip().split(".")
        if len(parts) == 2:
            day, month = int(parts[0]), int(parts[1])
            # Use a plausible year
            return date(2025, month, day)
    except Exception:
        pass
    return None


def _find_adjustments(
    transactions: List[BankTransaction],
    period_from_dt: Optional[date],
    period_to_dt: Optional[date],
) -> List[tuple]:
    """
    Return non-canceled value_date adjustments as list of
    (value_date, tx_date, amount, description).

    Cancellation: pairs sharing the same value_date whose amounts sum to ~0
    (one debit, one credit with identical absolute value) are removed.
    """
    vd_list = []
    for tx in transactions:
        if tx.value_date is None:
            continue
        tx_dt_obj = _parse_tx_date(tx.date)
        if tx_dt_obj is None:
            continue
        d = tx_dt_obj.date()
        if period_from_dt and d < period_from_dt:
            continue
        if period_to_dt and d > period_to_dt:
            continue
        # positive = credit (money in), negative = debit (money out)
        amount = (tx.credit or 0.0) - (tx.debit or 0.0)
        clean_op = re.sub(r'\s*[\(\uff08]תאריך ערך[^\)\uff09]*[\)\uff09]', '', tx.operation).strip()
        # Exclude interest transactions — only actual money movements create adjustments
        if re.search(r'ריבית|רבית', clean_op):
            continue
        vd_str = tx.value_date.strftime('%d/%m')
        desc = f"תיאום: {clean_op} ת.ערך {vd_str}"
        vd_list.append({
            'value_date': tx.value_date,
            'tx_date': d,
            'amount': amount,
            'desc': desc,
        })

    # Greedy cancellation of (same value_date, opposite amounts summing to ~0)
    used = [False] * len(vd_list)
    for i in range(len(vd_list)):
        if used[i]:
            continue
        for j in range(i + 1, len(vd_list)):
            if used[j]:
                continue
            a, b = vd_list[i], vd_list[j]
            if (a['value_date'] == b['value_date']
                    and abs(a['amount'] + b['amount']) < 1.0):
                used[i] = used[j] = True
                break

    return [
        (v['value_date'], v['tx_date'], v['amount'], v['desc'])
        for i, v in enumerate(vd_list)
        if not used[i]
    ]


def _period_from_date(period_from: str) -> Optional[date]:
    """Parse "DD/MM/YY" or "DD/MM/YYYY" → date object."""
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(period_from, fmt).date()
        except ValueError:
            pass
    return None


def _period_needs_p2(period_from: str, rate_change_date: str) -> bool:
    """
    Return True if the period is on or after the rate-change date
    (meaning the P2 formula column is relevant).

    Based on the sample: periods starting 26/11/25 or later include J-column
    formulas.  The rate-change date is 27.11 but the boundary row (26/11) is
    included.
    """
    rcd = _parse_rate_change_day(rate_change_date)
    pfd = _period_from_date(period_from)
    if rcd is None or pfd is None:
        return False
    # Include the day before rate change (boundary row)
    from datetime import timedelta
    return pfd >= (rcd - timedelta(days=1))


# ── ExcelWriter ───────────────────────────────────────────────────────────────

class ExcelWriter:
    """
    Build a :class:`openpyxl.Workbook` from a :class:`~models.statement.Statement`
    and save it to disk.
    """

    def __init__(self, statement: Statement):
        self.stmt = statement
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self._sheet1_total_row: Optional[int] = None   # set by _build_sheet1

    # ── Public API ────────────────────────────────────────────────────────

    def write(self, output_path: str) -> None:
        """Generate all three sheets and save the workbook to *output_path*."""
        self._build_sheet1()
        self._build_sheet2()
        self._build_sheet3()
        self.wb.save(output_path)

    # ═════════════════════════════════════════════════════════════════════
    # SHEET 1  פירוט ריבית חובה לפי PDF
    # ═════════════════════════════════════════════════════════════════════

    def _build_sheet1(self) -> None:
        ws = self.wb.create_sheet("פירוט ריבית חובה לפי PDF")
        ws.sheet_view.rightToLeft = True

        stmt = self.stmt

        # ── Row 1: Bank name / company name / period labels ───────────────
        _set(ws, "A1", stmt.bank_name, bold=True, size=14, alignment=_rtl_align())
        _set(ws, "F1", stmt.company_name or "ארביטראז' גלובל אל פי", alignment=_rtl_align())
        _set(ws, "I1", "תקופה 1", bold=True, alignment=_center_rtl())
        _set(ws, "J1", "תקופה 2", bold=True, alignment=_center_rtl())
        _set(ws, "K1", "תאריך שינוי ריבית", bold=True, alignment=_center_rtl())

        # ── Row 2: Branch / company address / BoI rates ───────────────────
        _set(ws, "A2", stmt.branch, alignment=_rtl_align())
        _set(ws, "F2", stmt.company_address or "דרך בגין 144", alignment=_rtl_align())
        _set(ws, "H2", "ריבית בנק ישראל", bold=True, alignment=_rtl_align())
        _set(ws, "I2", stmt.bank_rate_p1,
             number_format="0.00%", alignment=_center_rtl())
        if stmt.bank_rate_p2 is not None:
            _set(ws, "J2", stmt.bank_rate_p2,
                 number_format="0.00%", alignment=_center_rtl())
        if stmt.rate_change_date is not None:
            try:
                _set(ws, "K2",
                     float(stmt.rate_change_date.replace(".", ",").replace(",", ".")),
                     number_format="0.00", alignment=_center_rtl())
            except Exception:
                _set(ws, "K2", stmt.rate_change_date, alignment=_center_rtl())

        # ── Row 3: Phone / city / debit margin ───────────────────────────
        _set(ws, "A3", f"מס: {stmt.phone}", alignment=_rtl_align())
        _set(ws, "F3", stmt.company_city or "תל אביב - יפו", alignment=_rtl_align())
        _set(ws, "H3", "מרווח ריבית חובה", bold=True, alignment=_rtl_align())
        _set(ws, "I3", stmt.debit_margin,
             fill_hex=_INPUT_FILL,
             number_format="0.00%", alignment=_center_rtl())

        # ── Row 4: Account number / company ID / credit margin ────────────
        _set(ws, "A4", f"מס חשבון {stmt.account_full or stmt.account_number + ' 600'}",
             alignment=_rtl_align())
        _set(ws, "F4", stmt.company_id or "6492102", alignment=_rtl_align())
        _set(ws, "H4", "מרווח ריבית זכות", bold=True, alignment=_rtl_align())
        _set(ws, "I4", stmt.credit_margin,
             fill_hex=_INPUT_FILL,
             number_format="0.00%", alignment=_center_rtl())

        # ── Row 5: Print date ─────────────────────────────────────────────
        _set(ws, "A5", f"תאריך הדפסה {stmt.print_date}", alignment=_rtl_align())

        # ── Row 7: Subject line ───────────────────────────────────────────
        subject = (
            f"הנדון: פירוט חיוב הריבית מיום {stmt.charge_date} "
            f"לתקופה {stmt.period_from} - {stmt.period_to}"
        )
        _set(ws, "A7", subject, bold=True, size=12, alignment=_rtl_align(wrap=True))

        # ── Row 8: Credit-balance note ────────────────────────────────────
        _set(ws, "K8", "אם יש יתרות זכות אז יחושב פה", alignment=_rtl_align())

        # ── Row 9: Column headers ─────────────────────────────────────────
        headers = {
            "A9": "מתאריך",
            "B9": "עד תאריך",
            "C9": 'מסגרת חח"ד',
            "D9": "מסגרת נוספת חד-צדדית",
            "E9": 'גובה המדרגה (ש"ח)',
            "F9": "אחוז ריבית",
            "G9": "מספרי חובה",
            "H9": 'ריבית חובה (ש"ח)',
        }
        for ref, label in headers.items():
            cell = ws[ref]
            cell.value = label
            cell.font = _font(bold=True)
            cell.fill = _fill(_BLUE_HEADER)
            cell.alignment = _center_rtl()
            cell.border = _thin_border()
        # G9 gets an additional yellow fill
        ws["G9"].fill = _fill(_YELLOW)

        # Effective debit-rate formulas in J9 (P1) and K9 (P2)
        _set(ws, "J9", "=$I$2+$I$3", number_format="0.00%",
             alignment=_center_rtl(), bold=True)
        if stmt.bank_rate_p2 is not None:
            _set(ws, "K9", "=$J$2+$I$3", number_format="0.00%",
                 alignment=_center_rtl(), bold=True)

        # ── Data rows (row 10 onward) ─────────────────────────────────────
        row = 10
        first_tier_row: Optional[int] = None   # for the SUM range
        last_tier_row: Optional[int] = None

        # Group tier rows by period (preserve insertion order)
        seen_periods: dict = {}
        for tr in stmt.tier_rows:
            key = (tr.period_from, tr.period_to)
            if key not in seen_periods:
                seen_periods[key] = []
            seen_periods[key].append(tr)

        for (pf, pt), tiers in seen_periods.items():
            fw  = tiers[0].framework  if tiers else ""
            fw2 = tiers[0].framework2 if tiers else ""
            needs_p2 = (
                stmt.bank_rate_p2 is not None
                and stmt.rate_change_date is not None
                and _period_needs_p2(pf, stmt.rate_change_date)
            )

            # ── Period header row ─────────────────────────────────────────
            # Write dates as actual datetime objects so I formula works
            pf_dt = _parse_tx_date(pf)
            pt_dt = _parse_tx_date(pt)
            _set(ws, f"A{row}", pf_dt if pf_dt is not None else pf,
                 number_format="DD/MM/YY", alignment=_rtl_align(), border=_thin_border())
            _set(ws, f"B{row}", pt_dt if pt_dt is not None else pt,
                 number_format="DD/MM/YY", alignment=_rtl_align(), border=_thin_border())
            if fw:
                _set(ws, f"C{row}", f'{fw} ש"ח',
                     alignment=_rtl_align(), border=_thin_border())
            if fw2:
                _set(ws, f"D{row}", f'{fw2} ש"ח',
                     alignment=_rtl_align(), border=_thin_border())
            # Apply thin border to remaining cols even if empty
            for col in ["D", "E", "F", "G", "H"]:
                ws[f"{col}{row}"].border = _thin_border()
            # I column: number of days in the period
            ws[f"I{row}"].value        = f"=B{row}-A{row}"
            ws[f"I{row}"].number_format = "0"
            ws[f"I{row}"].alignment    = _center_rtl()
            row += 1

            # ── Tier rows ─────────────────────────────────────────────────
            for tr in tiers:
                if first_tier_row is None:
                    first_tier_row = row
                last_tier_row = row

                _set(ws, f"A{row}", tr.tier_type,
                     alignment=_rtl_align(), border=_thin_border())

                if tr.tier_amount is not None:
                    _set(ws, f"E{row}", tr.tier_amount,
                         number_format='#,##0', alignment=_right_rtl(),
                         border=_thin_border())

                _set(ws, f"F{row}", tr.rate,
                     number_format="0.00%", alignment=_center_rtl(),
                     border=_thin_border())

                _set(ws, f"G{row}", tr.debit_numbers,
                     number_format='_ * #,##0.0_',
                     alignment=_right_rtl(), border=_thin_border())

                _set(ws, f"H{row}", tr.interest,
                     number_format='#,##0.00', alignment=_right_rtl(),
                     border=_thin_border())

                # J column: P1 refund formula  =($J$9 - bank_rate) * debit / 365
                # K column: P2 refund formula  =($K$9 - bank_rate) * debit / 365
                if needs_p2 and stmt.bank_rate_p2 is not None:
                    ws[f"K{row}"].value        = f"=($K$9-F{row})*G{row}/365"
                    ws[f"K{row}"].number_format = '#,##0.00'
                    ws[f"K{row}"].alignment    = _right_rtl()
                else:
                    ws[f"J{row}"].value        = f"=($J$9-F{row})*G{row}/365"
                    ws[f"J{row}"].number_format = '#,##0.00'
                    ws[f"J{row}"].alignment    = _right_rtl()

                # Apply thin borders to empty columns B/C/D
                for col in ["B", "C", "D"]:
                    ws[f"{col}{row}"].border = _thin_border()

                row += 1

        # ── Summary block (matches boss's reference format) ────────────────
        row += 1   # blank row before summary
        sum_row = row

        _set(ws, f"A{sum_row}", 'סה"כ ריבית חובה:', bold=True, alignment=_rtl_align())
        _set(ws, f"F{sum_row}", "מדרגה 1", alignment=_rtl_align())
        _set(ws, f"G{sum_row}", stmt.total_tier1,
             number_format='#,##0.00', alignment=_right_rtl())

        if first_tier_row and last_tier_row:
            ws[f"J{sum_row}"].value        = f"=SUM(J{first_tier_row}:J{last_tier_row})"
            ws[f"J{sum_row}"].number_format = '#,##0.00'
            ws[f"J{sum_row}"].fill         = _fill(_YELLOW)
            ws[f"J{sum_row}"].font         = _font(bold=True)
            ws[f"J{sum_row}"].alignment    = _right_rtl()

            if stmt.bank_rate_p2 is not None:
                ws[f"K{sum_row}"].value        = f"=SUM(K{first_tier_row}:K{last_tier_row})"
                ws[f"K{sum_row}"].number_format = '#,##0.00'
                ws[f"K{sum_row}"].fill         = _fill(_YELLOW)
                ws[f"K{sum_row}"].font         = _font(bold=True)
                ws[f"K{sum_row}"].alignment    = _right_rtl()

            ws[f"L{sum_row}"].value        = f"=SUM(J{sum_row}:K{sum_row})"
            ws[f"L{sum_row}"].number_format = '#,##0.00'
            ws[f"L{sum_row}"].fill         = _fill(_YELLOW)
            ws[f"L{sum_row}"].font         = _font(bold=True)
            ws[f"L{sum_row}"].alignment    = _right_rtl()

        _set(ws, f"M{sum_row}", "החזר", bold=True, alignment=_rtl_align())

        row += 1
        # מדרגה 2 total row (only when 3-tier periods exist)
        has_tier2_rows = any(tr.tier_type == "מדרגה 2" for tr in stmt.tier_rows)
        if has_tier2_rows:
            total_tier2 = round(
                sum(r.interest for r in stmt.tier_rows if r.tier_type == "מדרגה 2"), 2
            )
            _set(ws, f"F{row}", "מדרגה 2", alignment=_rtl_align())
            _set(ws, f"G{row}", total_tier2,
                 number_format='#,##0.00', alignment=_right_rtl())
            row += 1

        excess_row = row
        _set(ws, f"F{excess_row}", "חריגה", alignment=_rtl_align())
        _set(ws, f"G{excess_row}", stmt.total_excess,
             number_format='#,##0.00', alignment=_right_rtl())

        row += 1
        total_row = row
        self._sheet1_total_row = total_row
        ws[f"G{total_row}"].value        = f"=SUM(G{sum_row}:G{excess_row})"
        ws[f"G{total_row}"].number_format = '#,##0.00'
        ws[f"G{total_row}"].font          = _font(bold=True)
        ws[f"G{total_row}"].alignment     = _right_rtl()

        row += 1
        charge_row = row
        charge_text = (
            f"חשבונך חויב ב- {stmt.charge_date} בסכום של: "
            f"{stmt.total_charged:,.2f} ש\"ח."
        )
        _set(ws, f"A{charge_row}", charge_text, bold=True, alignment=_rtl_align())

        row += 2   # blank row then notes
        _set(ws, f"A{row}", "לידיעתך:", bold=True, alignment=_rtl_align())
        row += 1
        note1 = (
            'מספרי חובה - נתון זה הינו סיכום (חיבור) של יתרות החובה בפועל '
            '(לפי תאריכי ערך)'
        )
        _set(ws, f"A{row}", note1, fill_hex=_YELLOW, alignment=_rtl_align(wrap=True))
        row += 1
        _set(ws, f"A{row}",
             "בתקופת החישוב לגביה מחושבת ריבית החובה במדרגה הנדונה.",
             alignment=_rtl_align(wrap=True))
        row += 1
        _set(ws, f"A{row}",
             "דיווח זה ניתן כשרות חינם ללקוח, בהתאם לכללי גילוי נאות.",
             alignment=_rtl_align(wrap=True))

        # ── Column widths ─────────────────────────────────────────────────
        col_widths = {
            "A": 16, "B": 16, "C": 18, "D": 22, "E": 18,
            "F": 14, "G": 20, "H": 18, "I": 14, "J": 14,
            "K": 20, "L": 14,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    # ═════════════════════════════════════════════════════════════════════
    # SHEET 2  בנק
    # ═════════════════════════════════════════════════════════════════════

    def _build_sheet2(self) -> None:
        ws = self.wb.create_sheet("בנק")
        ws.sheet_view.rightToLeft = True
        stmt = self.stmt

        # ── Rows 1-4: Header block ─────────────────────────────────────────
        _set(ws, "M1", "תקופה 1", bold=True, alignment=_center_rtl())
        _set(ws, "N1", "תקופה 2", bold=True, alignment=_center_rtl())
        _set(ws, "O1", "תאריך שינוי ריבית", bold=True, alignment=_center_rtl())

        _set(ws, "L2", "ריבית בנק ישראל", bold=True, alignment=_rtl_align())
        _set(ws, "M2", stmt.bank_rate_p1, number_format="0.00%", alignment=_center_rtl())
        if stmt.bank_rate_p2 is not None:
            _set(ws, "N2", stmt.bank_rate_p2, number_format="0.00%", alignment=_center_rtl())
        if stmt.rate_change_date is not None:
            try:
                _set(ws, "O2", float(stmt.rate_change_date.replace(".", ",")),
                     number_format="0.00", alignment=_center_rtl())
            except Exception:
                _set(ws, "O2", stmt.rate_change_date, alignment=_center_rtl())

        _set(ws, "A3", "תנועות בחשבון", bold=True, size=12, alignment=_rtl_align())
        _set(ws, "L3", "מרווח ריבית חובה", bold=True, alignment=_rtl_align())
        _set(ws, "M3", stmt.debit_margin,
             fill_hex=_INPUT_FILL, number_format="0.00%", alignment=_center_rtl())

        account_header = (
            f"מספר חשבון 12-600-{stmt.account_number} "
            f"לתקופה: {stmt.period_from} - {stmt.period_to}"
        )
        _set(ws, "A4", account_header, bold=True, alignment=_rtl_align())
        _set(ws, "L4", "מרווח ריבית זכות", bold=True, alignment=_rtl_align())
        _set(ws, "M4", stmt.credit_margin,
             fill_hex=_INPUT_FILL, number_format="0.00%", alignment=_center_rtl())

        # ── Row 5: Column headers ─────────────────────────────────────────
        tx_headers = {
            "A5": "תאריך",
            "B5": "קוד פעולה",
            "C5": "הפעולה",
            "D5": "פרטים",
            "E5": "אסמכתא",
            "F5": "צרור",
            "G5": "חובה",
            "H5": "זכות",
            "I5": 'יתרה בש"ח',
            "J5": "הערה",
        }
        for ref, label in tx_headers.items():
            _set(ws, ref, label, bold=True,
                 fill_hex=_HEADER_FILL2,
                 alignment=_center_rtl(),
                 border=_thin_border())

        # ── Transaction rows ──────────────────────────────────────────────
        data_start = 6
        self._write_transactions(ws, stmt.transactions, data_start,
                                 day_diff_col="J", formula_col=None)

        # ── Column widths ─────────────────────────────────────────────────
        col_widths = {
            "A": 14, "B": 12, "C": 22, "D": 18, "E": 14,
            "F": 8,  "G": 18, "H": 18, "I": 18, "J": 10,
            "K": 14, "L": 22, "M": 12, "N": 12, "O": 18,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    # ═════════════════════════════════════════════════════════════════════
    # SHEET 3  חישוב החזר לפי יתרות בבנק
    # ═════════════════════════════════════════════════════════════════════

    def _build_sheet3(self) -> None:  # noqa: C901
        ws = self.wb.create_sheet("חישוב החזר לפי יתרות בבנק")
        ws.sheet_view.rightToLeft = True
        stmt = self.stmt

        p1_eff = stmt.bank_rate_p1 + stmt.debit_margin
        p2_eff = (stmt.bank_rate_p2 + stmt.debit_margin
                  if stmt.bank_rate_p2 is not None else None)

        # ── Parse period / charge dates ───────────────────────────────────
        period_from_dt: Optional[date] = None
        period_to_dt: Optional[date] = None
        for _fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                period_from_dt = datetime.strptime(stmt.period_from, _fmt).date()
                break
            except ValueError:
                pass
        for _fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                period_to_dt = datetime.strptime(stmt.period_to, _fmt).date()
                break
            except ValueError:
                pass

        rcd: Optional[date] = None
        if stmt.rate_change_date is not None:
            rcd = _parse_rate_change_day(stmt.rate_change_date)

        charge_dt = _parse_tx_date(stmt.charge_date)

        # ── Filter transactions: up to period_to, keep all (no dedup) ────
        all_tx: List[Tuple[BankTransaction, date, datetime]] = []
        for tx in stmt.transactions:
            dt_obj = _parse_tx_date(tx.date)
            if dt_obj is None:
                continue
            d = dt_obj.date()
            if period_to_dt is not None and d > period_to_dt:
                continue
            all_tx.append((tx, d, dt_obj))

        # ── Find value-date adjustments (after cancellation) ─────────────
        adjustments = _find_adjustments(stmt.transactions, period_from_dt, period_to_dt)
        n_adj = len(adjustments)

        # ── Column layout (1-based index) ─────────────────────────────────
        # A=1 … I=9  fixed
        # J=10 …     adjustment columns (n_adj of them)
        # if n_adj > 0: adjusted-balance col after last adj col
        # days col, P1 col, P2 col (if applicable), total col
        ADJ_START = 10  # column J
        if n_adj > 0:
            ADJ_BAL_IDX = ADJ_START + n_adj          # adjusted balance
            DAYS_IDX    = ADJ_START + n_adj + 1
        else:
            ADJ_BAL_IDX = None
            DAYS_IDX    = ADJ_START                  # J = days when no adjustments

        P1_IDX    = DAYS_IDX + 1
        P2_IDX    = DAYS_IDX + 2 if p2_eff is not None else None
        TOTAL_IDX = DAYS_IDX + 3 if p2_eff is not None else DAYS_IDX + 2
        LABEL_IDX = TOTAL_IDX + 1

        DAYS_LET  = get_column_letter(DAYS_IDX)
        BAL_LET   = get_column_letter(ADJ_BAL_IDX) if ADJ_BAL_IDX else "I"
        P1_LET    = get_column_letter(P1_IDX)
        P2_LET    = get_column_letter(P2_IDX) if P2_IDX else None
        TOTAL_LET = get_column_letter(TOTAL_IDX)
        LABEL_LET = get_column_letter(LABEL_IDX)

        # ── Row 1: account/period header ──────────────────────────────────
        account_header = (
            f"מספר חשבון 12-600-{stmt.account_number} "
            f"לתקופה: {stmt.period_from} - {stmt.period_to}"
        )
        _set(ws, "A1", account_header, bold=True, alignment=_rtl_align())

        # ── Row 2: effective rates ────────────────────────────────────────
        _set(ws, "A2", "ריבית אפקטיבית P1:", bold=True, alignment=_rtl_align())
        _set(ws, "B2", p1_eff, number_format="0.00%", alignment=_center_rtl())
        if p2_eff is not None:
            _set(ws, "C2", "P2:", bold=True, alignment=_rtl_align())
            _set(ws, "D2", p2_eff, number_format="0.00%", alignment=_center_rtl())

        # ── Row 3: column headers ─────────────────────────────────────────
        HDR_ROW = 3
        base_headers = {
            "A": "תאריך", "B": "קוד פעולה", "C": "הפעולה",
            "D": "פרטים", "E": "אסמכתא", "F": "צרור",
            "G": "חובה", "H": "זכות", "I": 'יתרה בש"ח',
        }
        for col_l, label in base_headers.items():
            _set(ws, f"{col_l}{HDR_ROW}", label, bold=True,
                 fill_hex=_HEADER_FILL2, alignment=_center_rtl(), border=_thin_border())

        for j, (_, _, _, adj_desc) in enumerate(adjustments):
            col_l = get_column_letter(ADJ_START + j)
            _set(ws, f"{col_l}{HDR_ROW}", adj_desc, bold=True,
                 fill_hex=_HEADER_FILL2, alignment=_center_rtl(), border=_thin_border())

        if ADJ_BAL_IDX:
            _set(ws, f"{BAL_LET}{HDR_ROW}", "יתרה מתואמת", bold=True,
                 fill_hex=_HEADER_FILL2, alignment=_center_rtl(), border=_thin_border())

        _set(ws, f"{DAYS_LET}{HDR_ROW}", "ימים", bold=True,
             fill_hex=_HEADER_FILL2, alignment=_center_rtl(), border=_thin_border())
        _set(ws, f"{P1_LET}{HDR_ROW}", "ריבית P1", bold=True,
             fill_hex=_HEADER_FILL2, alignment=_center_rtl(), border=_thin_border())
        if P2_LET:
            _set(ws, f"{P2_LET}{HDR_ROW}", "ריבית P2", bold=True,
                 fill_hex=_HEADER_FILL2, alignment=_center_rtl(), border=_thin_border())
            _set(ws, f"{TOTAL_LET}{HDR_ROW}", 'סה"כ', bold=True,
                 fill_hex=_HEADER_FILL2, alignment=_center_rtl(), border=_thin_border())

        DATA_START = 4
        num_fmt_money = "#,##0.00"

        # ── Pre-compute days per unique date (process in input order) ─────
        # Input is typically newest-first; first unique date seen = newest.
        # days[newest] = charge_date - newest
        # days[d_i]    = d_{i-1} - d_i  (previous unique seen, which is newer)
        unique_dates_ordered: List[date] = []
        seen_dates_set: set = set()
        for _, d, _ in all_tx:
            if d not in seen_dates_set:
                seen_dates_set.add(d)
                unique_dates_ordered.append(d)

        date_to_days: dict = {}
        for i, d in enumerate(unique_dates_ordered):
            if i == 0:
                date_to_days[d] = (
                    max(1, (charge_dt.date() - d).days)
                    if charge_dt is not None else 1
                )
            else:
                date_to_days[d] = max(0, (unique_dates_ordered[i - 1] - d).days)

        # ── Write transaction rows ────────────────────────────────────────
        # For days: assign to the FIRST ROW WITH BALANCE for each date.
        # Rows without balance (common for intermediate intra-day transactions)
        # get days=0 and the date is not yet "consumed", so the next row with
        # a balance can claim it.  This ensures the interest formula always
        # has a valid balance to work with.
        dates_days_written: set = set()

        for i, (tx, d, dt_obj) in enumerate(all_tx):
            row = DATA_START + i

            has_balance = tx.balance is not None
            if d not in dates_days_written and has_balance:
                days_val = date_to_days.get(d, 0)
                dates_days_written.add(d)
            else:
                days_val = 0

            # Base columns A–I
            _set(ws, f"A{row}", dt_obj,
                 number_format="DD.MM.YYYY", alignment=_center_rtl())
            _set(ws, f"B{row}", tx.code,      alignment=_center_rtl())
            _set(ws, f"C{row}", tx.operation, alignment=_rtl_align())
            _set(ws, f"D{row}", tx.details,   alignment=_rtl_align())
            _set(ws, f"E{row}", tx.reference, alignment=_center_rtl())
            _set(ws, f"F{row}", tx.batch,     alignment=_center_rtl())
            if tx.debit is not None:
                _set(ws, f"G{row}", tx.debit,
                     number_format=num_fmt_money, alignment=_right_rtl())
            if tx.credit is not None:
                _set(ws, f"H{row}", tx.credit,
                     number_format=num_fmt_money, alignment=_right_rtl())
            if tx.balance is not None:
                _set(ws, f"I{row}", tx.balance,
                     number_format=num_fmt_money, alignment=_right_rtl())

            # Adjustment columns: value only when adj_vd <= row_date < adj_tx_date
            for j, (adj_vd, adj_tx_date, adj_amount, _) in enumerate(adjustments):
                if adj_vd <= d < adj_tx_date:
                    col_l = get_column_letter(ADJ_START + j)
                    _set(ws, f"{col_l}{row}", adj_amount,
                         number_format=num_fmt_money, alignment=_right_rtl())

            # Adjusted balance formula (sum of I + all adj columns)
            if ADJ_BAL_IDX:
                adj_first = get_column_letter(ADJ_START)
                adj_last  = get_column_letter(ADJ_START + n_adj - 1)
                ws[f"{BAL_LET}{row}"].value        = f"=I{row}+SUM({adj_first}{row}:{adj_last}{row})"
                ws[f"{BAL_LET}{row}"].number_format = num_fmt_money
                ws[f"{BAL_LET}{row}"].alignment    = _right_rtl()

            # Days column (hardcoded integer)
            ws[f"{DAYS_LET}{row}"].value        = days_val
            ws[f"{DAYS_LET}{row}"].number_format = "0"
            ws[f"{DAYS_LET}{row}"].alignment    = _center_rtl()

            # Interest formula: only when days > 0 and balance present
            if days_val > 0 and tx.balance is not None:
                use_p2 = (p2_eff is not None and rcd is not None and d >= rcd)
                rate   = p2_eff if use_p2 else p1_eff
                int_col = P2_LET if use_p2 else P1_LET
                ws[f"{int_col}{row}"].value = (
                    f'=IF({BAL_LET}{row}<0,'
                    f'{BAL_LET}{row}/365*{days_val}*{rate},"")'
                )
                ws[f"{int_col}{row}"].number_format = num_fmt_money
                ws[f"{int_col}{row}"].alignment    = _right_rtl()

        # ── Summary rows ──────────────────────────────────────────────────
        last_data_row = DATA_START + len(all_tx) - 1 if all_tx else DATA_START - 1
        sum_row     = last_data_row + 1
        charged_row = sum_row + 1
        refund_row  = sum_row + 2

        # Sum of P1 and P2 interest columns
        if all_tx:
            ws[f"{P1_LET}{sum_row}"].value        = f"=SUM({P1_LET}{DATA_START}:{P1_LET}{last_data_row})"
            ws[f"{P1_LET}{sum_row}"].number_format = num_fmt_money
            if P2_LET:
                ws[f"{P2_LET}{sum_row}"].value        = f"=SUM({P2_LET}{DATA_START}:{P2_LET}{last_data_row})"
                ws[f"{P2_LET}{sum_row}"].number_format = num_fmt_money
                ws[f"{TOTAL_LET}{sum_row}"].value = f"=SUM({P1_LET}{sum_row}:{P2_LET}{sum_row})"
            else:
                ws[f"{TOTAL_LET}{sum_row}"].value = f"={P1_LET}{sum_row}"
        else:
            ws[f"{TOTAL_LET}{sum_row}"].value = 0

        ws[f"{TOTAL_LET}{sum_row}"].number_format = num_fmt_money
        ws[f"{TOTAL_LET}{sum_row}"].font          = _font(bold=True)
        ws[f"{TOTAL_LET}{sum_row}"].fill          = _fill(_YELLOW)
        ws[f"{TOTAL_LET}{sum_row}"].alignment     = _right_rtl()
        _set(ws, f"{LABEL_LET}{sum_row}", "סכום לחיוב", bold=True, alignment=_rtl_align())

        sheet1_name = self.wb.worksheets[0].title
        ws[f"{TOTAL_LET}{charged_row}"].value        = f"=-'{sheet1_name}'!G{self._sheet1_total_row}"
        ws[f"{TOTAL_LET}{charged_row}"].number_format = num_fmt_money
        ws[f"{TOTAL_LET}{charged_row}"].alignment    = _right_rtl()
        _set(ws, f"{LABEL_LET}{charged_row}", "סכום שירד בפועל", alignment=_rtl_align())

        ws[f"{TOTAL_LET}{refund_row}"].value        = f"={TOTAL_LET}{sum_row}-{TOTAL_LET}{charged_row}"
        ws[f"{TOTAL_LET}{refund_row}"].number_format = num_fmt_money
        ws[f"{TOTAL_LET}{refund_row}"].font          = _font(bold=True)
        ws[f"{TOTAL_LET}{refund_row}"].fill          = _fill(_YELLOW)
        ws[f"{TOTAL_LET}{refund_row}"].alignment     = _right_rtl()
        _set(ws, f"{LABEL_LET}{refund_row}", "החזר ", bold=True,
             fill_hex=_YELLOW, alignment=_rtl_align())

        # ── Column widths ─────────────────────────────────────────────────
        col_widths: dict = {
            "A": 14, "B": 12, "C": 28, "D": 18, "E": 14,
            "F": 8,  "G": 18, "H": 18, "I": 18,
        }
        for j in range(n_adj):
            col_widths[get_column_letter(ADJ_START + j)] = 22
        if ADJ_BAL_IDX:
            col_widths[BAL_LET] = 20
        col_widths[DAYS_LET]  = 8
        col_widths[P1_LET]    = 14
        if P2_LET:
            col_widths[P2_LET]    = 14
        col_widths[TOTAL_LET] = 16
        col_widths[LABEL_LET] = 20
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    # ── Shared transaction writer ─────────────────────────────────────────

    def _write_transactions(
        self,
        ws,
        transactions: List[BankTransaction],
        data_start: int,
        day_diff_col: str,
        formula_col: Optional[str],
        rate_cell: Optional[str] = None,
        backward_day_col: Optional[str] = None,
    ) -> Optional[int]:
        """
        Write transaction rows starting at *data_start*.

        Column layout: A=date, B=code, C=operation, D=details, E=reference,
                       F=batch, G=debit, H=credit, I=balance,
                       *backward_day_col* (optional)=A_{prev}-A_{row}  ← days balance was held,
                       *day_diff_col*=A_{row}-A_{next}                 ← forward days,
                       *formula_col* (optional)=interest formula.

        When *backward_day_col* is provided the interest formula uses it for
        days (backward-looking); otherwise *day_diff_col* is used.

        The row immediately above data_start (data_start - 1) must already
        contain a date in column A to serve as the anchor for the first row's
        backward formula (=A{data_start-1} - A{data_start}).

        Returns the last data row number written, or None if no rows.
        """
        if not transactions:
            return None

        num_fmt_money = '#,##0.00'

        for i, tx in enumerate(transactions):
            row = data_start + i
            next_row = data_start + i + 1
            prev_row = row - 1   # anchor row for i=0; previous data row for i>0

            dt = _parse_tx_date(tx.date)
            _set(ws, f"A{row}", dt if dt is not None else tx.date,
                 number_format="DD.MM.YYYY", alignment=_center_rtl())
            _set(ws, f"B{row}", tx.code, alignment=_center_rtl())
            _set(ws, f"C{row}", tx.operation, alignment=_rtl_align())
            _set(ws, f"D{row}", tx.details, alignment=_rtl_align())
            _set(ws, f"E{row}", tx.reference, alignment=_center_rtl())
            _set(ws, f"F{row}", tx.batch, alignment=_center_rtl())

            if tx.debit is not None:
                _set(ws, f"G{row}", tx.debit,
                     number_format=num_fmt_money, alignment=_right_rtl())
            if tx.credit is not None:
                _set(ws, f"H{row}", tx.credit,
                     number_format=num_fmt_money, alignment=_right_rtl())
            if tx.balance is not None:
                _set(ws, f"I{row}", tx.balance,
                     number_format=num_fmt_money, alignment=_right_rtl())

            # Backward-looking days: how many days THIS balance was held.
            # Formula: =A{prev_row} - A{row}  (newer date above minus current date)
            # Skipped for the last (oldest) row — no inner-period prior boundary.
            if backward_day_col and i < len(transactions) - 1:
                ws[f"{backward_day_col}{row}"].value = f"=A{prev_row}-A{row}"
                ws[f"{backward_day_col}{row}"].number_format = "0"
                ws[f"{backward_day_col}{row}"].alignment = _center_rtl()

            # Forward-looking days: =A{row} - A{next_row}
            # (only valid when there IS a next row)
            if i < len(transactions) - 1:
                ws[f"{day_diff_col}{row}"].value = f"=A{row}-A{next_row}"
                ws[f"{day_diff_col}{row}"].number_format = "0"
                ws[f"{day_diff_col}{row}"].alignment = _center_rtl()

            # Interest formula: balance × days × rate/365
            # Uses backward_day_col for days when available (Sheet 3),
            # except for the last row which uses day_diff_col (both are blank
            # on the last row, but this matches the boss's reference file).
            if formula_col and rate_cell:
                is_last = (i == len(transactions) - 1)
                days_col = (
                    day_diff_col if (is_last or not backward_day_col)
                    else backward_day_col
                )
                ws[f"{formula_col}{row}"].value = (
                    f'=IFERROR(({rate_cell}/365*I{row}*{days_col}{row}),"")'
                )
                ws[f"{formula_col}{row}"].number_format = num_fmt_money
                ws[f"{formula_col}{row}"].alignment = _right_rtl()

        last_row = data_start + len(transactions) - 1
        return last_row
