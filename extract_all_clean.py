# -*- coding: utf-8 -*-
"""
Comprehensive extraction of 463560.xlsx and 463560 גלובל.pdf
Outputs clean Hebrew text to extract_all_clean_output.txt
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

OUTPUT_FILE = r'C:\Bank_Automation\extract_all_clean_output.txt'

lines = []

def w(s=''):
    lines.append(str(s))

# ============================================================
# EXCEL EXTRACTION
# ============================================================
w('=' * 80)
w('EXCEL FILE: 463560.xlsx')
w('=' * 80)

import openpyxl

wb = openpyxl.load_workbook(r'C:\Bank_Automation\463560.xlsx', data_only=False)
sheet_names = wb.sheetnames
w(f'\nSheet names: {sheet_names}')

for sheet_name in sheet_names:
    ws = wb[sheet_name]
    w()
    w('=' * 80)
    w(f"SHEET: '{sheet_name}'")
    w(f'  Dimensions: {ws.dimensions}')
    w(f'  Max row: {ws.max_row}, Max col: {ws.max_column}')

    # Merged cells
    w()
    w('  MERGED CELL RANGES:')
    merges = list(ws.merged_cells.ranges)
    if merges:
        for m in merges:
            w(f'    {m}')
    else:
        w('    (none)')

    w()
    w('  ROW-BY-ROW DATA (all non-empty cells, plus formula cells):')
    w()

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is None:
                continue
            coord = cell.coordinate
            dt = cell.data_type

            font = cell.font
            fill = cell.fill
            align = cell.alignment
            num_fmt = cell.number_format

            bold = font.bold if font else False
            italic = font.italic if font else False
            font_size = font.size if font else None
            font_name = font.name if font else None

            fg_color = None
            if fill and fill.fgColor:
                c = fill.fgColor
                if c.type == 'rgb':
                    fg_color = c.rgb
                elif c.type == 'indexed':
                    fg_color = f'indexed:{c.indexed}'
                elif c.type == 'theme':
                    fg_color = f'theme:{c.theme}'

            h_align = align.horizontal if align else None

            # Build concise line
            tags = []
            if bold:
                tags.append('BOLD')
            if italic:
                tags.append('ITALIC')
            if fg_color and fg_color not in ('00000000', 'indexed:64', None):
                tags.append(f'fill={fg_color}')
            if h_align:
                tags.append(f'align={h_align}')
            if num_fmt and num_fmt != 'General':
                tags.append(f'fmt={num_fmt}')

            tag_str = '  [' + ', '.join(tags) + ']' if tags else ''
            if dt == 'f':
                w(f'    {coord}: FORMULA = {val}{tag_str}')
            else:
                w(f'    {coord}: {repr(val)}{tag_str}')

    # Summary of formulas
    w()
    w('  FORMULAS SUMMARY:')
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                w(f'    {cell.coordinate}: {cell.value}')

w()
w('=' * 80)
w('END OF EXCEL')
w('=' * 80)

# ============================================================
# PDF EXTRACTION
# ============================================================
w()
w('=' * 80)
w('PDF FILE: 463560 גלובל.pdf')
w('=' * 80)

import pdfplumber

pdf_path = r'C:\Bank_Automation\463560 גלובל.pdf'

with pdfplumber.open(pdf_path) as pdf:
    w(f'\nTotal pages: {len(pdf.pages)}')
    w(f'Metadata: {pdf.metadata}')

    for page_num, page in enumerate(pdf.pages, 1):
        w()
        w('=' * 80)
        w(f'PAGE {page_num}')
        w(f'  Width: {page.width}, Height: {page.height}')

        # Extract text - try multiple methods
        w()
        w('  RAW TEXT (extract_text):')
        text = page.extract_text()
        if text:
            for line in text.split('\n'):
                w(f'    {line}')
        else:
            w('    (no text extracted)')

        # Try with layout=True for better RTL
        w()
        w('  TEXT WITH LAYOUT:')
        try:
            text_layout = page.extract_text(layout=True)
            if text_layout:
                for line in text_layout.split('\n'):
                    w(f'    {line}')
            else:
                w('    (no text)')
        except Exception as e:
            w(f'    Error: {e}')

        # Words with positions
        w()
        w('  WORDS WITH POSITIONS:')
        words = page.extract_words()
        w(f'  ({len(words)} words)')
        for wd in words:
            w(f"    x0={wd['x0']:.1f} y0={wd['top']:.1f} text={repr(wd['text'])}")

        # Tables
        w()
        w('  TABLES:')
        tables = page.extract_tables()
        if tables:
            for ti, table in enumerate(tables, 1):
                w(f'  Table {ti} ({len(table)} rows):')
                for ri, row in enumerate(table):
                    w(f'    Row {ri}: {row}')
        else:
            w('    (no tables found)')

w()
w('=' * 80)
w('END OF PDF')
w('=' * 80)

# Write output
with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'Done. Output written to {OUTPUT_FILE}')
print(f'Total lines: {len(lines)}')
