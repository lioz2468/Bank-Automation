import openpyxl

def cells(ws):
    d = {}
    for row in ws.iter_rows(min_row=1, max_row=35):
        for cell in row:
            val = cell.value
            if val is not None and val != '':
                d[cell.coordinate] = val
    return d

wb_out  = openpyxl.load_workbook('output2.xlsx', data_only=False)
wb_boss = openpyxl.load_workbook(r'C:\Users\lioz_\Downloads\זינו תבדוק בבקשה.xlsx', data_only=False)
out  = cells(wb_out.worksheets[2])
boss = cells(wb_boss.worksheets[2])
all_keys = sorted(set(out) | set(boss))
diffs = [(k, out.get(k), boss.get(k)) for k in all_keys if out.get(k) != boss.get(k)]
if diffs:
    print('DIFFERENCES:')
    for k, o, b in diffs:
        print(f'  {k}: output={repr(o)}  boss={repr(b)}')
else:
    print('PERFECT MATCH')
