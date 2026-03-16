import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('C:/Bank_Automation/463560.xlsx', data_only=False)

print("=" * 80)
print("EXCEL FILE: 463560.xlsx")
print("=" * 80)
print(f"\nSheet names: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("=" * 80)
    print(f"SHEET: '{sheet_name}'")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()

    # Merged cells
    print(f"  MERGED CELL RANGES:")
    if ws.merged_cells.ranges:
        for merge in ws.merged_cells.ranges:
            print(f"    {merge}")
    else:
        print("    (none)")
    print()

    # Column widths
    print(f"  COLUMN WIDTHS:")
    for col_letter, col_dim in ws.column_dimensions.items():
        print(f"    Col {col_letter}: width={col_dim.width}, hidden={col_dim.hidden}")
    print()

    # Row heights
    print(f"  ROW HEIGHTS:")
    for row_num, row_dim in ws.row_dimensions.items():
        print(f"    Row {row_num}: height={row_dim.height}, hidden={row_dim.hidden}")
    print()

    # All cells
    print(f"  ALL CELLS:")
    for row in ws.iter_rows():
        for cell in row:
            # Get font info
            font = cell.font
            font_info = {
                'bold': font.bold,
                'italic': font.italic,
                'size': font.size,
                'color': font.color.rgb if font.color and font.color.type == 'rgb' else (font.color.theme if font.color and font.color.type == 'theme' else str(font.color) if font.color else None),
                'name': font.name,
                'underline': font.underline,
            }

            # Get fill info
            fill = cell.fill
            fill_info = {
                'fill_type': fill.fill_type,
                'fgColor': fill.fgColor.rgb if fill.fgColor and fill.fgColor.type == 'rgb' else (str(fill.fgColor.theme) if fill.fgColor and fill.fgColor.type == 'theme' else str(fill.fgColor) if fill.fgColor else None),
                'bgColor': fill.bgColor.rgb if fill.bgColor and fill.bgColor.type == 'rgb' else (str(fill.bgColor.theme) if fill.bgColor and fill.bgColor.type == 'theme' else str(fill.bgColor) if fill.bgColor else None),
            }

            # Get alignment info
            align = cell.alignment
            align_info = {
                'horizontal': align.horizontal,
                'vertical': align.vertical,
                'wrap_text': align.wrap_text,
                'text_rotation': align.textRotation,
                'indent': align.indent,
                'reading_order': align.readingOrder,
            }

            # Get border info
            border = cell.border
            border_info = {
                'left': f"{border.left.border_style}/{border.left.color.rgb if border.left.color and border.left.color.type == 'rgb' else border.left.color}" if border.left else None,
                'right': f"{border.right.border_style}/{border.right.color.rgb if border.right.color and border.right.color.type == 'rgb' else border.right.color}" if border.right else None,
                'top': f"{border.top.border_style}/{border.top.color.rgb if border.top.color and border.top.color.type == 'rgb' else border.top.color}" if border.top else None,
                'bottom': f"{border.bottom.border_style}/{border.bottom.color.rgb if border.bottom.color and border.bottom.color.type == 'rgb' else border.bottom.color}" if border.bottom else None,
            }

            # Only print cells that have some content or formatting
            has_value = cell.value is not None
            has_formula = str(cell.value).startswith('=') if cell.value and isinstance(cell.value, str) else False
            has_format = cell.number_format != 'General'
            is_bold = font.bold
            has_fill = fill.fill_type not in (None, 'none')
            has_border = any([
                border.left and border.left.border_style,
                border.right and border.right.border_style,
                border.top and border.top.border_style,
                border.bottom and border.bottom.border_style,
            ])

            if has_value or has_format or is_bold or has_fill or has_border:
                print(f"    Cell {cell.coordinate}:")
                print(f"      value={repr(cell.value)}")
                print(f"      data_type={cell.data_type}")
                print(f"      number_format={cell.number_format}")
                print(f"      font={font_info}")
                print(f"      fill={fill_info}")
                print(f"      alignment={align_info}")
                print(f"      border={border_info}")

    print()

print("=" * 80)
print("END OF EXCEL EXTRACTION")
print("=" * 80)
